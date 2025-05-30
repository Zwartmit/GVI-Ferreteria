from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse
from django.views.decorators.cache import never_cache
from django.contrib.auth.decorators import login_required
from datetime import datetime, time
from app.models import *

################################################## Productos del Día ##################################################
@login_required
@never_cache
def export_productos_dia_excel(request, fecha_seleccionada=None):
    """
    Genera un reporte Excel que muestra los productos ingresados en una fecha específica.
    Si no se proporciona una fecha, usa la fecha actual.
    """
    # Si no se proporciona una fecha, usamos la fecha actual
    if not fecha_seleccionada:
        fecha_seleccionada = datetime.now().date()
    
    # Configuramos el inicio del día (00:00:00) y el final del día (23:59:59)
    inicio_del_dia = datetime.combine(fecha_seleccionada, time.min)
    fin_del_dia = datetime.combine(fecha_seleccionada, time.max)
    
    # Filtramos productos creados en la fecha seleccionada
    productos_del_dia = Producto.objects.filter(fecha_creacion__range=[inicio_del_dia, fin_del_dia])
    
    # Creamos el libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos del Día"
    
    # Estilos
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    green_fill = PatternFill(start_color="6B0606", end_color="6B0606", fill_type="solid")
    white_font = Font(color="FFFFFF")
    medium_border = Border(
        left=Side(style='medium'), 
        right=Side(style='medium'), 
        top=Side(style='medium'), 
        bottom=Side(style='medium')
    )
    
    # Configuramos anchos de columna
    column_width = 20
    for col in range(1, 10):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = column_width
    
    # Configuramos alturas de fila
    ws.row_dimensions[2].height = 38
    ws.row_dimensions[3].height = 23
    ws.row_dimensions[5].height = 20
    
    # Título y encabezado
    ws['E2'] = "GVI FERRETERIA"
    ws['E2'].font = Font(size=16, bold=True)
    
    ws.merge_cells('B2:I2')
    ws['B2'].alignment = center_alignment
    ws['B2'].border = medium_border
    
    ws.merge_cells('B3:I3')
    ws['B3'] = f"Reporte de Productos Ingresados el {fecha_seleccionada.strftime('%d/%m/%Y')}"
    ws['B3'].font = Font(size=14, bold=True)
    ws['B3'].alignment = center_alignment
    ws['B3'].border = medium_border
    
    # Fecha del reporte
    ws.merge_cells('B4:I4')
    fecha_generacion = datetime.now().strftime("%d/%m/%Y")
    ws['B4'] = f"Fecha de generación: {fecha_generacion}"
    ws['B4'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws['B4'].border = medium_border
    
    # Encabezados de columnas
    headers = ['ID', 'Producto', 'Cantidad', 'Costo', 'Precio Venta', 'Estado', 'Categoría', 'Presentación', 'Hora']
    
    for col_num, header in enumerate(headers, 2):
        cell = ws.cell(row=5, column=col_num)
        cell.value = header
        cell.fill = green_fill
        cell.font = white_font
        cell.alignment = center_alignment
        cell.border = medium_border
    
    # Agregamos los productos al reporte
    if productos_del_dia.exists():
        for row_num, producto in enumerate(productos_del_dia, 6):
            ws.cell(row=row_num, column=2, value=producto.id)
            ws.cell(row=row_num, column=3, value=producto.producto)
            ws.cell(row=row_num, column=4, value=producto.cantidad)
            ws.cell(row=row_num, column=5, value=producto.valor)
            ws.cell(row=row_num, column=6, value=producto.precio_venta)
            ws.cell(row=row_num, column=7, value='Activo' if producto.estado else 'Inactivo')
            ws.cell(row=row_num, column=8, value=producto.id_categoria.categoria)
            ws.cell(row=row_num, column=9, value=producto.id_presentacion.presentacion)
            ws.cell(row=row_num, column=10, value=producto.fecha_creacion.strftime('%H:%M:%S'))
            
            for col_num in range(2, 11):
                cell = ws.cell(row=row_num, column=col_num)
                cell.alignment = center_alignment
                cell.border = medium_border
    else:
        # Si no hay productos en la fecha seleccionada, agregamos un mensaje
        ws.merge_cells('B6:I6')
        ws['B6'] = f"No se encontraron productos ingresados el {fecha_seleccionada.strftime('%d/%m/%Y')}"
        ws['B6'].font = bold_font
        ws['B6'].alignment = center_alignment
    
    # Configuramos el nombre del archivo y devolvemos la respuesta
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Productos_del_dia_{fecha_seleccionada.strftime("%d-%m-%Y")}.xlsx'
    wb.save(response)
    return response

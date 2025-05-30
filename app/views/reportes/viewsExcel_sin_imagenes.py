from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse
from django.views.decorators.cache import never_cache
from django.contrib.auth.decorators import login_required
from datetime import datetime
from app.models import *

################################################## Productos ##################################################
@login_required
@never_cache
def export_productos_excel(request, fecha_inicio=None, fecha_fin=None):
    productos = Producto.objects.all()
    
    if fecha_inicio and fecha_fin:
        productos = productos.filter(fecha_creacion__range=[fecha_inicio, fecha_fin])

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de productos"

    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    green_fill = PatternFill(start_color="6B0606", end_color="6B0606", fill_type="solid")
    white_font = Font(color="FFFFFF")
    medium_border = Border(left=Side(style='medium'), 
                         right=Side(style='medium'), 
                         top=Side(style='medium'), 
                         bottom=Side(style='medium'))

    column_width = 20  
    for col in range(2, 10): 
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = column_width

    ws.row_dimensions[2].height = 38 
    ws.row_dimensions[3].height = 23  
    ws.row_dimensions[5].height = 20

    # En lugar de usar imágenes, usamos texto para el encabezado
    ws['E2'] = "GVI FERRETERIA"
    ws['E2'].font = Font(size=16, bold=True)
    
    ws.merge_cells('B2:I2')  
    ws['B2'].alignment = center_alignment
    ws['B2'].border = medium_border

    ws.merge_cells('B3:I3')
    ws['B3'] = "Reporte de Productos"
    ws['B3'].font = Font(size=14, bold=True)
    ws['B3'].alignment = center_alignment
    ws['B3'].border = medium_border
    for col in range(3, 10):
        ws.cell(row=3, column=col).border = medium_border

    ws.merge_cells('B4:I4')
    fecha = datetime.now().strftime("%d/%m/%Y")
    ws['B4'] = f"Fecha: {fecha}"
    ws['B4'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws['B4'].border = medium_border
    for col in range(3, 10):
        ws.cell(row=4, column=col).border = medium_border

    headers = ['ID', 'Producto', 'Cantidad', 'Valor', 'Estado', 'Categoría', 'Presentación']
    for col_num, header in enumerate(headers, 2):
        cell = ws.cell(row=5, column=col_num)
        cell.value = header
        cell.fill = green_fill
        cell.font = white_font
        cell.alignment = center_alignment
        cell.border = medium_border

    for row_num, producto in enumerate(productos, 6):  
        ws.cell(row=row_num, column=2, value=producto.id)
        ws.cell(row=row_num, column=3, value=producto.producto)
        ws.cell(row=row_num, column=4, value=producto.cantidad)
        ws.cell(row=row_num, column=5, value=producto.valor)
        ws.cell(row=row_num, column=6, value='Activo' if producto.estado else 'Inactivo')
        ws.cell(row=row_num, column=7, value=producto.id_categoria.categoria)
        ws.cell(row=row_num, column=8, value=f"{producto.id_presentacion.presentacion}") 
        
        for col_num in range(2, 10): 
            cell = ws.cell(row=row_num, column=col_num)
            cell.alignment = center_alignment
            cell.border = medium_border

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Reporte_de_productos.xlsx'
    wb.save(response)
    return response

################################################## Administradores ##################################################
@login_required
@never_cache
def export_administradores_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de administradores"

    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    green_fill = PatternFill(start_color="6B0606", end_color="6B0606", fill_type="solid")
    white_font = Font(color="FFFFFF")
    medium_border = Border(left=Side(style='medium'), 
                           right=Side(style='medium'), 
                           top=Side(style='medium'), 
                           bottom=Side(style='medium'))

    column_width = 20  
    for col in range(2, 8): 
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = column_width

    ws.row_dimensions[2].height = 38 
    ws.row_dimensions[3].height = 23  
    ws.row_dimensions[5].height = 20

    # En lugar de usar imágenes, usamos texto para el encabezado
    ws['D2'] = "GVI FERRETERIA"
    ws['D2'].font = Font(size=16, bold=True)
    
    ws.merge_cells('B2:G2')
    ws['B2'].alignment = center_alignment
    ws['B2'].border = medium_border
    for col in range(3, 8):
        ws.cell(row=2, column=col).border = medium_border

    ws.merge_cells('B3:G3')
    ws['B3'] = "Reporte de Administradores"
    ws['B3'].font = Font(size=14, bold=True)
    ws['B3'].alignment = center_alignment
    ws['B3'].border = medium_border
    for col in range(3, 8):
        ws.cell(row=3, column=col).border = medium_border

    ws.merge_cells('B4:G4')
    fecha = datetime.now().strftime("%d/%m/%Y")
    ws['B4'] = f"Fecha: {fecha}"
    ws['B4'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws['B4'].border = medium_border
    for col in range(3, 8):
        ws.cell(row=4, column=col).border = medium_border

    headers = ['ID', 'Nombre', 'Tipo de documento', '# de documento', 'Email', 'Teléfono']
    for col_num, header in enumerate(headers, 2):
        cell = ws.cell(row=5, column=col_num)
        cell.value = header
        cell.fill = green_fill
        cell.font = white_font
        cell.alignment = center_alignment
        cell.border = medium_border

    administradores = Administrador.objects.all()
    for row_num, administrador in enumerate(administradores, 6):  
        ws.cell(row=row_num, column=2, value=administrador.id)
        ws.cell(row=row_num, column=3, value=administrador.nombre)
        ws.cell(row=row_num, column=4, value=administrador.tipo_documento)
        ws.cell(row=row_num, column=5, value=administrador.numero_documento)
        ws.cell(row=row_num, column=6, value=administrador.user.email)
        ws.cell(row=row_num, column=7, value=administrador.telefono)

        for col_num in range(2, 8):
            cell = ws.cell(row=row_num, column=col_num)
            cell.alignment = center_alignment
            cell.border = medium_border

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Reporte de administradores.xlsx'
    wb.save(response)
    return response

################################################## Operadores ##################################################
@login_required
@never_cache
def export_operadores_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de operadores"

    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    green_fill = PatternFill(start_color="6B0606", end_color="6B0606", fill_type="solid")
    white_font = Font(color="FFFFFF")
    medium_border = Border(left=Side(style='medium'), 
                           right=Side(style='medium'), 
                           top=Side(style='medium'), 
                           bottom=Side(style='medium'))

    column_width = 20  
    for col in range(2, 8): 
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = column_width

    ws.row_dimensions[2].height = 38 
    ws.row_dimensions[3].height = 23  
    ws.row_dimensions[5].height = 20

    # En lugar de usar imágenes, usamos texto para el encabezado
    ws['D2'] = "GVI FERRETERIA"
    ws['D2'].font = Font(size=16, bold=True)
    
    ws.merge_cells('B2:G2')
    ws['B2'].alignment = center_alignment
    ws['B2'].border = medium_border
    for col in range(3, 8):
        ws.cell(row=2, column=col).border = medium_border

    ws.merge_cells('B3:G3')
    ws['B3'] = "Reporte de Operadores"
    ws['B3'].font = Font(size=14, bold=True)
    ws['B3'].alignment = center_alignment
    ws['B3'].border = medium_border
    for col in range(3, 8):
        ws.cell(row=3, column=col).border = medium_border

    ws.merge_cells('B4:G4')
    fecha = datetime.now().strftime("%d/%m/%Y")
    ws['B4'] = f"Fecha: {fecha}"
    ws['B4'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws['B4'].border = medium_border
    for col in range(3, 8):
        ws.cell(row=4, column=col).border = medium_border

    headers = ['ID', 'Nombre', 'Tipo de documento', '# de documento', 'Email', 'Teléfono']
    for col_num, header in enumerate(headers, 2):
        cell = ws.cell(row=5, column=col_num)
        cell.value = header
        cell.fill = green_fill
        cell.font = white_font
        cell.alignment = center_alignment
        cell.border = medium_border

    operadores = Operador.objects.all()
    for row_num, operador in enumerate(operadores, 6):  
        ws.cell(row=row_num, column=2, value=operador.id)
        ws.cell(row=row_num, column=3, value=operador.nombre)
        ws.cell(row=row_num, column=4, value=operador.tipo_documento)
        ws.cell(row=row_num, column=5, value=operador.numero_documento)
        ws.cell(row=row_num, column=6, value=operador.user.email)
        ws.cell(row=row_num, column=7, value=operador.telefono)

        for col_num in range(2, 8):
            cell = ws.cell(row=row_num, column=col_num)
            cell.alignment = center_alignment
            cell.border = medium_border

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Reporte de operadores.xlsx'
    wb.save(response)
    return response

################################################## Ventas ##################################################
@login_required
@never_cache
def export_ventas_excel(request, fecha_inicio=None, fecha_fin=None):
    # Importamos datetime para obtener la fecha actual
    from datetime import datetime, time
    
    # Obtenemos la fecha actual
    hoy = datetime.now().date()
    
    # Configuramos el inicio del día (00:00:00) y el final del día (23:59:59)
    inicio_del_dia = datetime.combine(hoy, time.min)
    fin_del_dia = datetime.combine(hoy, time.max)
    
    # Filtramos solamente las ventas del día actual
    ventas = Venta.objects.filter(fecha_venta__range=[inicio_del_dia, fin_del_dia])

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de ventas"

    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    green_fill = PatternFill(start_color="6B0606", end_color="6B0606", fill_type="solid")
    white_font = Font(color="FFFFFF")
    medium_border = Border(left=Side(style='medium'), 
                           right=Side(style='medium'), 
                           top=Side(style='medium'), 
                           bottom=Side(style='medium'))

    column_width = 20  
    for col in range(2, 8): 
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = column_width

    ws.row_dimensions[2].height = 38 
    ws.row_dimensions[3].height = 23  
    ws.row_dimensions[5].height = 20

    # En lugar de usar imágenes, usamos texto para el encabezado
    ws['D2'] = "GVI FERRETERIA"
    ws['D2'].font = Font(size=16, bold=True)
    
    ws.merge_cells('B2:G2')
    ws['B2'].alignment = center_alignment
    ws['B2'].border = medium_border
    for col in range(3, 8):
        ws.cell(row=2, column=col).border = medium_border

    ws.merge_cells('B3:G3')
    ws['B3'] = "Reporte de Ventas del Día"
    ws['B3'].font = Font(size=14, bold=True)
    ws['B3'].alignment = center_alignment
    ws['B3'].border = medium_border
    for col in range(3, 8):
        ws.cell(row=3, column=col).border = medium_border

    ws.merge_cells('B4:G4')
    fecha = datetime.now().strftime("%d/%m/%Y")
    ws['B4'] = f"Fecha: {fecha}"
    ws['B4'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws['B4'].border = medium_border
    for col in range(3, 8):
        ws.cell(row=4, column=col).border = medium_border

    # Cambiamos los encabezados para reflejar los campos reales del modelo
    headers = ['ID', 'Total Venta', 'Método Pago', 'Dinero Recibido', 'Cambio', 'Fecha']
    for col_num, header in enumerate(headers, 2):
        cell = ws.cell(row=5, column=col_num)
        cell.value = header
        cell.fill = green_fill
        cell.font = white_font
        cell.alignment = center_alignment
        cell.border = medium_border

    for row_num, venta in enumerate(ventas, 6):
        ws.cell(row=row_num, column=2, value=venta.id)
        ws.cell(row=row_num, column=3, value=venta.total_venta)
        # Convertimos el código del método de pago a su versión legible
        if venta.metodo_pago == 'EF':
            metodo_pago_display = 'Efectivo'
        elif venta.metodo_pago == 'TF':
            metodo_pago_display = 'Transferencia'
        else:
            metodo_pago_display = venta.metodo_pago
        ws.cell(row=row_num, column=4, value=metodo_pago_display)
        ws.cell(row=row_num, column=5, value=venta.dinero_recibido)
        ws.cell(row=row_num, column=6, value=venta.cambio)
        ws.cell(row=row_num, column=7, value=venta.fecha_venta.strftime("%d/%m/%Y %H:%M") if venta.fecha_venta else "")

        for col_num in range(2, 8):
            cell = ws.cell(row=row_num, column=col_num)
            cell.alignment = center_alignment
            cell.border = medium_border

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Reporte de ventas.xlsx'
    wb.save(response)
    return response

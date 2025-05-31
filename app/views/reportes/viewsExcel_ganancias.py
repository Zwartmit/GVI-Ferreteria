from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse
from django.views.decorators.cache import never_cache
from django.contrib.auth.decorators import login_required
from datetime import datetime, time, timedelta
from app.models import *
from decimal import Decimal
from django.db.models import Sum, F

################################################## Ganancias Diarias ##################################################
@login_required
@never_cache
def export_ganancias_diarias_excel(request):
    """
    Genera un reporte Excel con las ganancias del día actual,
    mostrando costos, precios de venta y ganancias por producto.
    """
    # Intentamos obtener fechas del formulario
    fecha_inicio = request.POST.get('fecha_inicio') or request.GET.get('fecha_inicio')
    fecha_fin = request.POST.get('fecha_fin') or request.GET.get('fecha_fin')

    hoy = datetime.now().date()
    titulo_fecha = hoy.strftime('%d/%m/%Y')
    if fecha_inicio and fecha_fin:
        try:
            inicio_del_dia = datetime.combine(datetime.strptime(fecha_inicio, '%Y-%m-%d'), time.min)
            fin_del_dia = datetime.combine(datetime.strptime(fecha_fin, '%Y-%m-%d'), time.max)
            # Para el título, muestra el rango
            titulo_fecha = f"{inicio_del_dia.strftime('%d/%m/%Y')} - {fin_del_dia.strftime('%d/%m/%Y')}"
        except Exception:
            inicio_del_dia = datetime.combine(hoy, time.min)
            fin_del_dia = datetime.combine(hoy, time.max)
            titulo_fecha = hoy.strftime('%d/%m/%Y')
    else:
        inicio_del_dia = datetime.combine(hoy, time.min)
        fin_del_dia = datetime.combine(hoy, time.max)
        titulo_fecha = hoy.strftime('%d/%m/%Y')

    # Obtenemos todas las ventas del día o rango
    ventas_hoy = Venta.objects.filter(fecha_venta__range=[inicio_del_dia, fin_del_dia])

    # Creamos el libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Ganancias Diarias"

    # Estilos
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    green_fill = PatternFill(start_color="6B0606", end_color="6B0606", fill_type="solid")
    white_font = Font(color="FFFFFF")
    medium_border = Border(
        left=Side(style='medium'), 
        right=Side(style='medium'), 
        top=Side(style='medium'), 
        bottom=Side(style='medium')
    )

    # Configuramos anchos de columna
    column_width = 18
    for col in range(1, 12):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = column_width

    # Configuramos alturas de fila
    ws.row_dimensions[2].height = 38
    ws.row_dimensions[3].height = 23
    ws.row_dimensions[5].height = 20

    # Título y encabezado
    ws['D2'] = "GVI FERRETERIA"
    ws['D2'].font = Font(size=16, bold=True)

    ws.merge_cells('B2:J2')
    ws['B2'].alignment = center_alignment
    ws['B2'].border = medium_border

    ws.merge_cells('B3:J3')
    ws['B3'] = f"Reporte de Ganancias Diarias - {titulo_fecha}"
    ws['B3'].font = Font(size=14, bold=True)
    ws['B3'].alignment = center_alignment
    ws['B3'].border = medium_border

    row_num = 5
    costo_total = Decimal('0.0')
    venta_total = Decimal('0.0')
    ganancia_total = Decimal('0.0')

    for venta in ventas_hoy:
        # Escribir encabezado de la venta
        headers_venta = ["ID Venta", "Fecha", "Total Venta", "Método Pago", "Dinero Recibido", "Cambio"]
        for col_num, header in enumerate(headers_venta, 2):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = header
            cell.fill = green_fill
            cell.font = white_font
            cell.alignment = center_alignment
            cell.border = medium_border
        row_num += 1
        # Escribir datos de la venta
        datos_venta = [
            venta.id,
            venta.fecha_venta.strftime("%d/%m/%Y %H:%M") if venta.fecha_venta else "",
            float(venta.total_venta),
            venta.get_metodo_pago_display() if hasattr(venta, 'get_metodo_pago_display') else venta.metodo_pago,
            float(venta.dinero_recibido),
            float(venta.cambio)
        ]
        for col_num, dato in enumerate(datos_venta, 2):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = dato
            cell.alignment = center_alignment
            cell.border = medium_border
        row_num += 1
        # Encabezado de productos
        headers_prod = ["Producto", "Cantidad", "Costo Unitario", "Precio Venta", "Costo Total", "Venta Total", "Ganancia"]
        for col_num, header in enumerate(headers_prod, 2):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = header
            cell.fill = green_fill
            cell.font = white_font
            cell.alignment = center_alignment
            cell.border = medium_border
        row_num += 1
        # Productos de la venta
        detalles = Detalle_venta.objects.filter(id_venta=venta)
        for detalle in detalles:
            nombre_producto = detalle.nombre_producto
            cantidad = detalle.cantidad_producto
            precio_venta = detalle.valor_unitario
            subtotal = detalle.subtotal_venta
            if detalle.id_producto:
                costo_unitario = detalle.id_producto.valor
                costo = costo_unitario * cantidad
                ganancia = subtotal - costo
            else:
                costo_unitario = Decimal('0.0')
                costo = Decimal('0.0')
                ganancia = Decimal('0.0')
            costo_total += costo
            venta_total += subtotal
            ganancia_total += ganancia
            datos_prod = [
                nombre_producto,
                cantidad,
                costo_unitario,
                precio_venta,
                costo,
                subtotal,
                ganancia
            ]
            for col_num, dato in enumerate(datos_prod, 2):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = dato
                cell.alignment = center_alignment
                cell.border = medium_border
            row_num += 1
        # Fila vacía de separación
        row_num += 1

    # Agregamos el resumen total
    ws.cell(row=row_num + 1, column=2, value="TOTALES:")
    ws.cell(row=row_num + 1, column=2).font = bold_font
    ws.cell(row=row_num + 1, column=6, value=costo_total)
    ws.cell(row=row_num + 1, column=7, value=venta_total)
    ws.cell(row=row_num + 1, column=8, value=ganancia_total)

    for col_num in [2, 6, 7, 8]:
        cell = ws.cell(row=row_num + 1, column=col_num)
        cell.font = bold_font
        cell.alignment = center_alignment
        cell.border = medium_border

    # Configuramos el nombre del archivo y devolvemos la respuesta
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Ganancias_Diarias.xlsx'
    wb.save(response)
    return response

################################################## Ganancias Mensuales ##################################################
@login_required
@never_cache
def export_ganancias_mensuales_excel(request):
    """
    Genera un reporte Excel con las ganancias del mes actual,
    mostrando costos, precios de venta y ganancias por día.
    """
    # Intentamos obtener fechas del formulario
    fecha_inicio = request.POST.get('fecha_inicio') or request.GET.get('fecha_inicio')
    fecha_fin = request.POST.get('fecha_fin') or request.GET.get('fecha_fin')

    hoy = datetime.now().date()
    nombre_mes = {
        1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril', 5: 'Mayo', 6: 'Junio',
        7: 'Julio', 8: 'Agosto', 9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
    }
    titulo_fecha = f"{nombre_mes[hoy.month]} {hoy.year}"
    if fecha_inicio and fecha_fin:
        try:
            inicio_mes = datetime.combine(datetime.strptime(fecha_inicio, '%Y-%m-%d'), time.min)
            fin_mes = datetime.combine(datetime.strptime(fecha_fin, '%Y-%m-%d'), time.max)
            # Para el título, muestra el rango
            titulo_fecha = f"{inicio_mes.strftime('%d/%m/%Y')} - {fin_mes.strftime('%d/%m/%Y')}"
        except Exception:
            primer_dia_mes = hoy.replace(day=1)
            if hoy.month == 12:
                ultimo_dia_mes = hoy.replace(day=31)
            else:
                ultimo_dia_mes = hoy.replace(month=hoy.month+1, day=1) - timedelta(days=1)
            inicio_mes = datetime.combine(primer_dia_mes, time.min)
            fin_mes = datetime.combine(ultimo_dia_mes, time.max)
            titulo_fecha = f"{nombre_mes[hoy.month]} {hoy.year}"
    else:
        primer_dia_mes = hoy.replace(day=1)
        if hoy.month == 12:
            ultimo_dia_mes = hoy.replace(day=31)
        else:
            ultimo_dia_mes = hoy.replace(month=hoy.month+1, day=1) - timedelta(days=1)
        inicio_mes = datetime.combine(primer_dia_mes, time.min)
        fin_mes = datetime.combine(ultimo_dia_mes, time.max)
        titulo_fecha = f"{nombre_mes[hoy.month]} {hoy.year}"

    # Obtenemos todas las ventas del mes o rango
    ventas_mes = Venta.objects.filter(fecha_venta__range=[inicio_mes, fin_mes])
    
    # Obtenemos los detalles de todas las ventas del mes
    detalles_ventas = Detalle_venta.objects.filter(id_venta__in=ventas_mes)
    
    # Creamos un diccionario para almacenar los totales por día
    resumen_diario = {}
    
    # Calculamos las ganancias por día
    for detalle in detalles_ventas:
        fecha_venta = detalle.id_venta.fecha_venta.date()
        
        if detalle.id_producto:
            costo_unitario = detalle.id_producto.valor
            costo = costo_unitario * detalle.cantidad_producto
        else:
            costo = Decimal('0.0')
            
        venta = detalle.subtotal_venta
        ganancia = venta - costo
        
        if fecha_venta not in resumen_diario:
            resumen_diario[fecha_venta] = {
                'costo_total': costo,
                'venta_total': venta,
                'ganancia_total': ganancia
            }
        else:
            resumen_diario[fecha_venta]['costo_total'] += costo
            resumen_diario[fecha_venta]['venta_total'] += venta
            resumen_diario[fecha_venta]['ganancia_total'] += ganancia
    
    # Calculamos los totales del mes
    costo_total_mes = sum(dia['costo_total'] for dia in resumen_diario.values())
    venta_total_mes = sum(dia['venta_total'] for dia in resumen_diario.values())
    ganancia_total_mes = sum(dia['ganancia_total'] for dia in resumen_diario.values())
    
    # Creamos el libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Ganancias Mensuales"
    
    # Estilos
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    green_fill = PatternFill(start_color="6B0606", end_color="6B0606", fill_type="solid")
    white_font = Font(color="FFFFFF")
    medium_border = Border(
        left=Side(style='medium'), 
        right=Side(style='medium'), 
        top=Side(style='medium'), 
        bottom=Side(style='medium')
    )
    
    # Configuramos anchos de columna
    column_width = 20
    for col in range(1, 6):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = column_width
    
    # Configuramos alturas de fila
    ws.row_dimensions[2].height = 38
    ws.row_dimensions[3].height = 23
    ws.row_dimensions[5].height = 20
    
    # Título y encabezado
    ws['D2'] = "GVI FERRETERIA"
    ws['D2'].font = Font(size=16, bold=True)

    ws.merge_cells('B2:J2')
    ws['B2'].alignment = center_alignment
    ws['B2'].border = medium_border

    ws.merge_cells('B3:J3')
    ws['B3'] = f"Reporte de Ganancias Mensuales - {nombre_mes[hoy.month]} {hoy.year}"
    ws['B3'].font = Font(size=14, bold=True)
    ws['B3'].alignment = center_alignment
    ws['B3'].border = medium_border

    row_num = 5
    costo_total = Decimal('0.0')
    venta_total = Decimal('0.0')
    ganancia_total = Decimal('0.0')

    ventas_mes = Venta.objects.filter(fecha_venta__range=[inicio_mes, fin_mes])
    for venta in ventas_mes:
        # Escribir encabezado de la venta
        headers_venta = ["ID Venta", "Fecha", "Total Venta", "Método Pago", "Dinero Recibido", "Cambio"]
        for col_num, header in enumerate(headers_venta, 2):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = header
            cell.fill = green_fill
            cell.font = white_font
            cell.alignment = center_alignment
            cell.border = medium_border
        row_num += 1
        # Escribir datos de la venta
        datos_venta = [
            venta.id,
            venta.fecha_venta.strftime("%d/%m/%Y %H:%M") if venta.fecha_venta else "",
            float(venta.total_venta),
            venta.get_metodo_pago_display() if hasattr(venta, 'get_metodo_pago_display') else venta.metodo_pago,
            float(venta.dinero_recibido),
            float(venta.cambio)
        ]
        for col_num, dato in enumerate(datos_venta, 2):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = dato
            cell.alignment = center_alignment
            cell.border = medium_border
        row_num += 1
        # Encabezado de productos
        headers_prod = ["Producto", "Cantidad", "Costo Unitario", "Precio Venta", "Costo Total", "Venta Total", "Ganancia"]
        for col_num, header in enumerate(headers_prod, 2):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = header
            cell.fill = green_fill
            cell.font = white_font
            cell.alignment = center_alignment
            cell.border = medium_border
        row_num += 1
        # Productos de la venta
        detalles = Detalle_venta.objects.filter(id_venta=venta)
        for detalle in detalles:
            nombre_producto = detalle.nombre_producto
            cantidad = detalle.cantidad_producto
            precio_venta = detalle.valor_unitario
            subtotal = detalle.subtotal_venta
            if detalle.id_producto:
                costo_unitario = detalle.id_producto.valor
                costo = costo_unitario * cantidad
                ganancia = subtotal - costo
            else:
                costo_unitario = Decimal('0.0')
                costo = Decimal('0.0')
                ganancia = Decimal('0.0')
            costo_total += costo
            venta_total += subtotal
            ganancia_total += ganancia
            datos_prod = [
                nombre_producto,
                cantidad,
                costo_unitario,
                precio_venta,
                costo,
                subtotal,
                ganancia
            ]
            for col_num, dato in enumerate(datos_prod, 2):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = dato
                cell.alignment = center_alignment
                cell.border = medium_border
            row_num += 1
        # Fila vacía de separación
        row_num += 1

    # Agregamos el resumen total
    ws.cell(row=row_num + 1, column=2, value="TOTALES:")
    ws.cell(row=row_num + 1, column=2).font = bold_font
    ws.cell(row=row_num + 1, column=6, value=costo_total)
    ws.cell(row=row_num + 1, column=7, value=venta_total)
    ws.cell(row=row_num + 1, column=8, value=ganancia_total)

    for col_num in [2, 6, 7, 8]:
        cell = ws.cell(row=row_num + 1, column=col_num)
        cell.font = bold_font
        cell.alignment = center_alignment
        cell.border = medium_border

    # Configuramos el nombre del archivo y devolvemos la respuesta
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Ganancias_Mensuales.xlsx'
    wb.save(response)
    return response
